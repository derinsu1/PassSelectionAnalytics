from __future__ import annotations

import csv
import io
import json
import math
import re
from dataclasses import dataclass
from datetime import UTC, datetime
from typing import Any, Literal

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from pass_selection.api.contracts import AppliedReviewFilters, PlayerStatsFilters
from pass_selection.api.repository import WorkbenchRepository

ExportScope = Literal["review_explorer", "pass_inspector", "player_directory", "player_profile", "player_comparison"]
ExportFormat = Literal["xlsx", "json", "csv"]


class ExportAnnotation(BaseModel):
    decision_id: str | None = None
    player_id: int | None = None
    status: str = "unreviewed"
    note: str = ""
    created_at: str | None = None
    updated_at: str | None = None
    author: str | None = None


class ExportRequest(BaseModel):
    scope: ExportScope
    format: ExportFormat
    review_filters: AppliedReviewFilters = Field(default_factory=AppliedReviewFilters)
    review_sort_by: str = "minute"
    review_sort_direction: Literal["asc", "desc"] = "asc"
    decision_id: str | None = None
    player_id: int | None = None
    comparison_player_id: int | None = None
    player_filters: PlayerStatsFilters | None = None
    timeline_window: int = Field(default=30, ge=1, le=180)
    annotations: list[ExportAnnotation] = Field(default_factory=list)


@dataclass(frozen=True)
class ExportFile:
    content: bytes
    filename: str
    media_type: str


def _plain(value: Any) -> Any:
    if isinstance(value, BaseModel):
        return _plain(value.model_dump())
    if isinstance(value, dict):
        return {str(key): _plain(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_plain(item) for item in value]
    if isinstance(value, float) and not math.isfinite(value):
        return None
    return value


def _flatten(value: Any, prefix: str = "") -> dict[str, Any]:
    source = _plain(value)
    if not isinstance(source, dict):
        return {prefix or "value": source}
    flattened: dict[str, Any] = {}
    for key, item in source.items():
        name = f"{prefix}_{key}" if prefix else key
        if isinstance(item, dict):
            flattened.update(_flatten(item, name))
        elif isinstance(item, list):
            flattened[name] = json.dumps(item, ensure_ascii=False, separators=(",", ":"))
        else:
            flattened[name] = item
    return flattened


_REVIEW_EXPORT_FIELDS = (
    "decision_id", "match_id", "match_name", "period", "frame", "match_clock", "team_id", "team_name",
    "passer_id", "passer_name", "passer_origin_third", "passer_origin_side", "selected_receiver_id",
    "selected_receiver_name", "selected_open_xt_rank", "selected_open_xt", "selected_open_xt_delta",
    "highest_open_xt_receiver_id", "highest_open_xt_receiver_name", "highest_open_xt", "highest_open_xt_delta",
    "local_open_xt_margin", "selected_local_xpass", "selected_availability_score",
    "selected_pass_viability_score", "selected_pass_viability_rank", "best_pass_viability_receiver_id",
    "best_pass_viability_receiver_name", "best_pass_viability_score", "pass_viability_gap", "pass_outcome",
)
_OPTION_EXPORT_FIELDS = (
    "option_id", "receiver_id", "receiver_name", "is_selected", "is_highest_pvi", "tracking_quality",
    "same_frame", "local_xpass", "pass_viability",
)


def _allow(value: Any, fields: tuple[str, ...]) -> dict[str, Any]:
    plain = _plain(value)
    return {key: plain[key] for key in fields if key in plain}


def _local_review(value: Any) -> dict[str, Any]:
    return _allow(value, _REVIEW_EXPORT_FIELDS)


def _local_option(value: Any) -> dict[str, Any]:
    return _allow(value, _OPTION_EXPORT_FIELDS)


def _annotation_rows(annotations: list[ExportAnnotation], *, decision_ids: set[str] | None = None, player_ids: set[int] | None = None) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for annotation in annotations:
        if decision_ids is not None and annotation.decision_id not in decision_ids:
            continue
        if player_ids is not None and annotation.player_id not in player_ids:
            continue
        if annotation.decision_id is None and annotation.player_id is None:
            continue
        rows.append({
            "content_type": "analyst_annotation",
            "canonical_analytics": False,
            "author": annotation.author or "local analyst",
            "decision_id": annotation.decision_id,
            "player_id": annotation.player_id,
            "manual_status": annotation.status,
            "status": annotation.status,
            "note": annotation.note,
            "analyst_rationale": annotation.note,
            "created_at": annotation.created_at or annotation.updated_at,
            "updated_at": annotation.updated_at or annotation.created_at,
            "annotation_semantics": "User-authored review material; not canonical analytics and does not alter model classifications.",
        })
    return rows


class ExportService:
    """Builds side-effect-free exports from immutable artifacts plus request-local notes."""

    def __init__(self, repository: WorkbenchRepository):
        self.repository = repository

    def render(self, request: ExportRequest) -> ExportFile:
        payload = self.payload(request)
        if request.format == "json":
            json_payload = {
                key: value
                for key, value in payload.items()
                if key not in {"worksheets", "csv_rows", "json"}
            }
            json_payload["data"] = payload["json"]
            return ExportFile(
                content=(json.dumps(json_payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n").encode(),
                filename=f"{payload['filename_stem']}.json",
                media_type="application/json",
            )
        if request.format == "csv":
            rows = payload.get("csv_rows")
            if rows is None:
                raise ValueError("CSV is available only for a single flat table on this tab.")
            return ExportFile(
                content=self._csv(rows),
                filename=f"{payload['filename_stem']}.csv",
                media_type="text/csv; charset=utf-8",
            )
        return ExportFile(
            content=self._xlsx(payload),
            filename=f"{payload['filename_stem']}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def payload(self, request: ExportRequest) -> dict[str, Any]:
        exported_at = datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")
        metadata = _plain(self.repository.metadata())
        common = {
            "exported_at": exported_at,
            "application": metadata["application"],
            "analytical_artifact_revision": metadata["analytical_artifact_revision"],
            "data_source": metadata["data_source"],
            "annotation_semantics": "Analyst annotations are browser-local review material. They are not canonical analytics and never change outcomes or published local metrics.",
            "missing_value_convention": "Blank cells represent unavailable scalar values; zero is never used to stand in for missing data.",
            "metric_timing": "Local metrics use the actual pass frame and compare tracked teammates where the metric defines a comparison universe.",
        }
        if request.scope == "review_explorer":
            return self._review_payload(request, common)
        if request.scope == "pass_inspector":
            return self._inspector_payload(request, common)
        if request.scope in {"player_directory", "player_profile", "player_comparison"}:
            return self._player_payload(request, common)
        raise ValueError(f"Unsupported export scope: {request.scope}")

    def _review_payload(self, request: ExportRequest, common: dict[str, Any]) -> dict[str, Any]:
        response = self.repository.list_reviews(
            request.review_filters, 1, 10_000, request.review_sort_by, request.review_sort_direction
        )
        local_reviews = [_local_review(item) for item in response.items]
        rows = [_flatten(item) for item in local_reviews]
        notes = _annotation_rows(request.annotations, decision_ids={row["decision_id"] for row in rows})
        notes_by_id = {str(row["decision_id"]): row for row in notes if row["decision_id"]}
        for row in rows:
            annotation = notes_by_id.get(str(row["decision_id"]))
            row["annotation_status"] = annotation["status"] if annotation else "unreviewed"
            row["short_analyst_note"] = annotation["note"] if annotation else None
        return {
            **common,
            "scope": "Review Explorer",
            "filename_stem": f"review-explorer_{datetime.now(UTC).date().isoformat()}",
            "export_scope": f"All {response.total} decisions matching active filters; pagination is excluded.",
            "active_filters": _plain({**response.applied_filters.model_dump(), "sort_by": response.sort_by, "sort_direction": response.sort_direction}),
            "worksheets": {
                "Decisions": rows,
                "Annotations": notes,
                "Active Filters": [_flatten({**response.applied_filters.model_dump(), "sort_by": response.sort_by, "sort_direction": response.sort_direction})],
                "Metric Definitions": [{"definition": value} for value in self._metric_definitions()],
            },
            "csv_rows": rows,
            "json": {"reviews": local_reviews, "annotations": notes, "active_filters": _plain(response.applied_filters)},
        }

    def _inspector_payload(self, request: ExportRequest, common: dict[str, Any]) -> dict[str, Any]:
        if not request.decision_id:
            raise ValueError("A decision_id is required for a Pass Inspector export.")
        detail = self.repository.decision_details(request.decision_id)
        playback = self.repository.playback(request.decision_id, request.timeline_window)
        notes = _annotation_rows(request.annotations, decision_ids={request.decision_id})
        if not notes:
            notes = [{
                "content_type": "analyst_annotation", "canonical_analytics": False,
                "author": "local analyst", "decision_id": request.decision_id, "player_id": None,
                "manual_status": "unreviewed", "status": "unreviewed", "note": "No analyst annotation recorded",
                "analyst_rationale": "No analyst annotation recorded", "created_at": None, "updated_at": None,
                "annotation_semantics": "User-authored review material; not canonical analytics and does not alter model classifications.",
            }]
        return {
            **common,
            "scope": "Pass Inspector",
            "filename_stem": f"pass-inspector_{request.decision_id.replace(':', '-').replace('_', '-')}_frame-{detail.provider_pass_frame}",
            "export_scope": f"Decision {request.decision_id}; source-frame playback window ±{request.timeline_window} frames.",
            "active_filters": {"decision_id": request.decision_id, "timeline_window": request.timeline_window},
            "worksheets": {
                "Decision Summary": [_flatten(_local_review(detail.summary))],
                "Receiver Options": [_flatten(_local_option(item)) for item in detail.options],
                "Timeline": [_flatten(item) for item in playback.frames],
                "Current Frame Players": [_flatten(player) for frame in playback.frames if frame.frame_number == detail.provider_pass_frame for player in frame.players],
                "Annotation and Notes": notes,
                "Metric Definitions": [{"definition": value} for value in detail.metric_definitions],
            },
            "json": {"decision": {"summary": _local_review(detail.summary), "selected_receiver": _local_option(detail.selected_receiver) if detail.selected_receiver else None, "highest_pvi_receiver": _local_option(detail.highest_pvi_receiver) if detail.highest_pvi_receiver else None, "options": [_local_option(item) for item in detail.options], "actual_pass_frame": detail.provider_pass_frame, "metric_definitions": detail.metric_definitions}, "playback": {"frames": _plain(playback.frames), "actual_pass_frame": detail.provider_pass_frame}, "annotations": notes},
        }

    def _player_payload(self, request: ExportRequest, common: dict[str, Any]) -> dict[str, Any]:
        filters = request.player_filters or PlayerStatsFilters()
        if request.scope == "player_profile":
            if request.player_id is None:
                raise ValueError("A player_id is required for a player profile export.")
            profile = self.repository.player_stats_detail(request.player_id, filters.match_id, filters.team_id)
            notes = _annotation_rows(request.annotations, player_ids={request.player_id})
            if not notes:
                notes = [{"content_type": "analyst_annotation", "canonical_analytics": False, "author": "local analyst", "decision_id": None, "player_id": request.player_id, "manual_status": "unreviewed", "status": "unreviewed", "note": "No analyst annotation recorded", "analyst_rationale": "No analyst annotation recorded", "created_at": None, "updated_at": None, "annotation_semantics": "User-authored review material; not canonical analytics and does not alter model classifications."}]
            return {
                **common, "scope": "Player Analysis profile", "filename_stem": f"player-profile_{request.player_id}",
                "export_scope": f"Player {request.player_id}; fixed available-match profile.", "active_filters": _plain(filters),
                "worksheets": {
                    "Player Summary": [_flatten(profile.model_dump(exclude={"match_breakdown", "percentiles"}))],
                    "Player Metrics": [_flatten(profile.passing) | {f"receiving_{key}": value for key, value in _flatten(profile.receiving).items()}],
                    "Match Breakdown": [_flatten(item) for item in profile.match_breakdown],
                    "Cohort Comparison": [_flatten(item) for item in profile.percentiles],
                    "Annotations and Notes": notes,
                    "Metric Definitions": [{"definition": value} for value in self._metric_definitions()],
                },
                "json": {"player": _plain(profile), "annotations": notes, "active_filters": _plain(filters)},
            }
        if request.scope == "player_comparison":
            if request.player_id is None or request.comparison_player_id is None:
                raise ValueError("Both player IDs are required for a player comparison export.")
            if request.player_id == request.comparison_player_id:
                raise ValueError("A player comparison requires two distinct players.")
            anchor = self.repository.player_stats_detail(request.player_id, None, None)
            comparison = self.repository.player_stats_detail(request.comparison_player_id, None, None)
            anchor_metrics = {metric.metric: metric for metric in anchor.percentiles}
            comparison_metrics = {metric.metric: metric for metric in comparison.percentiles}
            metric_rows = []
            for code, anchor_metric in anchor_metrics.items():
                comparison_metric = comparison_metrics[code]
                delta = (
                    None
                    if anchor_metric.value is None or comparison_metric.value is None
                    else comparison_metric.value - anchor_metric.value
                )
                metric_rows.append({
                    "metric": code,
                    "label": anchor_metric.label,
                    "direction": anchor_metric.direction,
                    "value_format": anchor_metric.value_format,
                    "anchor_value": anchor_metric.value,
                    "anchor_percentile": anchor_metric.percentile,
                    "anchor_peer_count": anchor_metric.peer_count,
                    "comparison_value": comparison_metric.value,
                    "comparison_percentile": comparison_metric.percentile,
                    "comparison_peer_count": comparison_metric.peer_count,
                    "comparison_minus_anchor": delta,
                })
            player_ids = {request.player_id, request.comparison_player_id}
            notes = _annotation_rows(request.annotations, player_ids=player_ids)
            return {
                **common,
                "scope": "Player Analysis comparison",
                "filename_stem": f"player-comparison_{request.player_id}-vs-{request.comparison_player_id}",
                "export_scope": f"Fixed available-match comparison: player {request.player_id} versus player {request.comparison_player_id}.",
                "active_filters": {"anchor_player_id": request.player_id, "comparison_player_id": request.comparison_player_id},
                "worksheets": {
                    "Comparison Summary": [
                        _flatten({"role": "Viewed player", **anchor.model_dump(exclude={"match_breakdown", "percentiles"})}),
                        _flatten({"role": "Comparison player", **comparison.model_dump(exclude={"match_breakdown", "percentiles"})}),
                    ],
                    "Metric Comparison": metric_rows,
                    "Viewed Player Matches": [_flatten(item) for item in anchor.match_breakdown],
                    "Comparison Player Matches": [_flatten(item) for item in comparison.match_breakdown],
                    "Annotations and Notes": notes,
                    "Metric Definitions": [{"definition": value} for value in self._metric_definitions()],
                },
                "csv_rows": metric_rows,
                "json": {
                    "viewed_player": _plain(anchor),
                    "comparison_player": _plain(comparison),
                    "metrics": metric_rows,
                    "annotations": notes,
                },
            }
        directory = self.repository.list_player_stats(
            match_id=filters.match_id, team_id=filters.team_id, search=filters.search,
            hide_under_60=filters.hide_under_60, position=filters.position,
            percentile_metric=filters.percentile_metric, min_percentile=filters.min_percentile,
            max_percentile=filters.max_percentile, sort_by="minutes", sort_direction="desc",
        )
        rows = [_flatten(item) for item in directory.items]
        directory_notes = _annotation_rows(
            request.annotations, player_ids={int(item.player_id) for item in directory.items}
        )
        return {
            **common, "scope": "Player Analysis directory", "filename_stem": f"player-analysis_{datetime.now(UTC).date().isoformat()}",
            "export_scope": f"All {directory.total} players matching active membership filters.", "active_filters": _plain(directory.applied_filters),
            "worksheets": {
                "Players": rows, "Annotations": directory_notes,
                "Active Filters": [_flatten(directory.applied_filters)],
                "Cohort Definitions": [{"definition": "Percentiles are fixed against 60+ minute players in the same primary position across all available local matches."}],
                "Metric Definitions": [{"definition": value} for value in self._metric_definitions()],
            },
            "csv_rows": rows,
            "json": {"players": [_plain(item) for item in directory.items], "annotations": directory_notes, "active_filters": _plain(directory.applied_filters)},
        }

    @staticmethod
    def _metric_definitions() -> list[str]:
        return [
            "Same-frame Local xT v1 is a location-only value proxy at the authoritative actual pass frame.",
            "Local xPass v1 is conditional on deliberately attempting a direct pass to the named teammate.",
            "Availability v1 is a deterministic lane/interception proxy, not a probability.",
            "PVI is a deterministic safety/value composite, not a pass-selection verdict or player rating.",
            "PVI v2 is 55% Local xPass and 45% bounded actual-frame delta-xT utility; Availability is a visible local diagnostic, not a PVI input or probability.",
        ]

    @staticmethod
    def _csv(rows: list[dict[str, Any]]) -> bytes:
        fields = sorted({key for row in rows for key in row})
        stream = io.StringIO(newline="")
        writer = csv.DictWriter(stream, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: "" if row.get(key) is None else row.get(key) for key in fields})
        return stream.getvalue().encode("utf-8-sig")

    def _xlsx(self, payload: dict[str, Any]) -> bytes:
        workbook = Workbook(write_only=True)
        readme_rows = []
        for key in ("scope", "export_scope", "active_filters", "metric_timing", "missing_value_convention", "annotation_semantics", "application", "analytical_artifact_revision", "data_source", "exported_at"):
            value = payload.get(key)
            readme_rows.append({"Field": key.replace("_", " ").title(), "Value": json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value})
        names = [*payload["worksheets"], "Notes", "Export Metadata"]
        readme_rows.append({"Field": "Worksheet descriptions", "Value": f"{', '.join(names)}. Each worksheet contains its named structured analytical dataset. Blank scalar cells mean unavailable values."})
        self._write_table(workbook, "README", readme_rows)
        for name, rows in payload["worksheets"].items():
            self._write_table(workbook, name, rows)
        self._write_table(workbook, "Notes", payload["worksheets"].get("Annotations") or payload["worksheets"].get("Annotation and Notes") or payload["worksheets"].get("Annotations and Notes") or [])
        self._write_table(workbook, "Export Metadata", [{key: value for key, value in payload.items() if key not in {"worksheets", "json", "csv_rows"}}])
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _write_table(self, workbook: Workbook, name: str, rows: list[dict[str, Any]]) -> None:
        safe = re.sub(r"[\\/*?:\[\]]", "-", name)[:31] or "Sheet"
        if safe in workbook.sheetnames:
            return
        sheet = workbook.create_sheet(safe)
        fields = list(dict.fromkeys(key for row in rows for key in row))
        if not fields:
            fields = ["note"]
            rows = [{"note": "No analyst annotation recorded"}] if "note" in safe.casefold() or "annotation" in safe.casefold() else [{"note": "No rows available for this export scope."}]
        sheet.freeze_panes = "A2"
        sheet.append([self._header_cell(sheet, field) for field in fields])
        for row in rows:
            sheet.append([self._data_cell(sheet, self._excel_value(row.get(field))) for field in fields])
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{len(rows) + 1}"
        for index, field in enumerate(fields, start=1):
            sample = [str(row.get(field) or "") for row in rows[:100]]
            sheet.column_dimensions[get_column_letter(index)].width = min(48, max(12, len(field) + 2, max(map(len, sample), default=0) + 2))

    @staticmethod
    def _excel_value(value: Any) -> Any:
        value = _plain(value)
        if value is None:
            return None
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
        return value

    @staticmethod
    def _data_cell(sheet: Any, value: Any) -> Any:
        if not isinstance(value, str) or len(value) <= 40:
            return value
        cell = WriteOnlyCell(sheet, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        return cell

    @staticmethod
    def _header_cell(sheet: Any, value: str) -> WriteOnlyCell:
        cell = WriteOnlyCell(sheet, value=value)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        return cell
